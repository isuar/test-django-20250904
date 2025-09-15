from django.core.management.base import BaseCommand
from django.conf import settings
from pathlib import Path
from openpyxl import load_workbook

from product.models import ProductCategory, RawMaterial

class Command(BaseCommand):
    help = "Import ProductCategory and RawMaterial from data_files/*.xlsx using openpyxl (no pandas)."

    def handle(self, *args, **kwargs):
        data_dir = Path(settings.BASE_DIR).parent / 'data_files'
        imported_cats = imported_rm = 0

        # --- Product Categories ---
        cat_path = data_dir / 'product_category.xlsx'
        if cat_path.exists():
            wb = load_workbook(cat_path, read_only=True, data_only=True)
            ws = wb.active
            # assume first row is header; find 'name' column
            headers = {str(c.value).strip().lower(): idx for idx, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1))[0:])}
            name_idx = headers.get('name', 0)  # fallback to first col
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[name_idx]
                name = (str(val).strip() if val is not None else "")
                if name:
                    _, created = ProductCategory.objects.get_or_create(name=name)
                    imported_cats += int(created)
            self.stdout.write(self.style.SUCCESS(f"Product categories imported/created: {imported_cats}"))
        else:
            self.stdout.write(self.style.WARNING("product_category.xlsx not found."))

        # --- Raw Materials ---
        rm_path = data_dir / 'raw_material.xlsx'
        if rm_path.exists():
            wb = load_workbook(rm_path, read_only=True, data_only=True)
            ws = wb.active
            headers = {str(c.value).strip().lower(): idx for idx, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1))[0:])}
            name_idx = headers.get('name', 0)
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[name_idx]
                name = (str(val).strip() if val is not None else "")
                if name:
                    _, created = RawMaterial.objects.get_or_create(name=name)
                    imported_rm += int(created)
            self.stdout.write(self.style.SUCCESS(f"Raw materials imported/created: {imported_rm}"))
        else:
            self.stdout.write(self.style.WARNING("raw_material.xlsx not found."))
